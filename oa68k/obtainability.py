"""GENUINELY_UNOBTAINABLE must be EARNED. This module is the only thing that can grant it.

THE RULE, and it is the whole point of the file:

    A 404 from an API is NOT a demonstration that a document does not exist.

A failed URL probe measures OUR REACH. To say a datum is genuinely unobtainable you
must consult a REGISTER THAT ENUMERATES THE POPULATION and show the item is absent
from the register's own rows. That is the register answering, not the network.

THE MODEL INSTANCE (why this module has the shape it has):

    EMA publishes `medicines-output-medicines-report_en.xlsx` -- every medicine it
    has assessed centrally, one row each. Zero rows for four iron products. An EPAR
    exists ONLY for a centrally authorised medicine, so "absent from this register"
    IS the register's own answer: there is no EPAR to fetch. That is earned.
    `GET <guessed-epar-url> -> 404` is not.

FOUR REQUIREMENTS, all four, or the verdict is refused:

  1. ENUMERATION NAMED     -- which register, its URL, what ONE ROW means.
  2. RETRIEVAL DATE + HASH -- a claim about a register is a claim about a VERSION.
  3. POSITIVE CONTROL      -- a key KNOWN to be in the register, found in the SAME
                              BYTES, BEFORE any negative is accepted. An enumeration
                              that can only answer "absent" is not a check.
  4. ABSENCE FROM ROWS     -- the query key is absent from the register's own rows.
                              Evidence of kind HTTP_STATUS is rejected outright.

The positive control is run FIRST and against the SAME BYTES the negative is read
from ("a verifier must search the SAME bytes it showed").

Run:  python obtainability.py --selftest      # plants both ways, no network
      python obtainability.py --ema --query "Ferric carboxymaltose"
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from enum import Enum
from typing import Callable, Iterable

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


class EvidenceKind(str, Enum):
    """What KIND of thing is being offered as evidence of absence."""
    ENUMERATION_ROWS = "enumeration_rows"   # the register's own rows -- the only valid kind
    HTTP_STATUS = "http_status"             # a 404/403/500 -- REJECTED, measures our reach
    SEARCH_MISS = "search_miss"             # a query returned nothing -- REJECTED, measures the query
    ASSERTION = "assertion"                 # someone said so -- REJECTED


VALID_EVIDENCE = {EvidenceKind.ENUMERATION_ROWS}


def normalise_key(k: str) -> str:
    return " ".join(str(k).strip().lower().split())


@dataclass
class Enumeration:
    """A register that enumerates a population, pinned to a version.

    `keys` is the full set of normalised row keys. `row_meaning` states what ONE ROW
    is -- without it nobody can tell whether absence means anything at all.
    """
    name: str
    url: str
    row_meaning: str
    retrieved_utc: str
    sha256: str
    n_rows: int
    keys: frozenset
    positive_control_key: str = ""
    positive_control_passed: bool = False
    positive_control_note: str = ""
    absence_licenses: str = ""

    @staticmethod
    def from_bytes(name: str, url: str, row_meaning: str, absence_licenses: str,
                   payload: bytes, extract_keys: Callable[[bytes], Iterable[str]],
                   positive_control_key: str) -> "Enumeration":
        """Build from the RAW BYTES. Hash, rows and control all come from one object.

        The positive control is evaluated HERE, against THESE bytes, before the
        Enumeration is usable -- so it cannot be skipped or run against a re-read.
        """
        digest = hashlib.sha256(payload).hexdigest()
        keys = frozenset(normalise_key(k) for k in extract_keys(payload) if str(k).strip())
        pc = normalise_key(positive_control_key)
        passed = bool(pc) and pc in keys
        return Enumeration(
            name=name, url=url, row_meaning=row_meaning,
            retrieved_utc=datetime.now(timezone.utc).isoformat(timespec="seconds"),
            sha256=digest, n_rows=len(keys), keys=keys,
            positive_control_key=positive_control_key,
            positive_control_passed=passed,
            positive_control_note=(
                "positive control " + repr(positive_control_key) + " "
                + ("FOUND" if passed else "NOT FOUND")
                + " among the " + str(len(keys)) + " keys of sha256:" + digest[:12]),
            absence_licenses=absence_licenses)

    def contains(self, key: str) -> bool:
        return normalise_key(key) in self.keys

    def summary(self) -> dict:
        d = asdict(self)
        d["keys"] = "<" + str(len(self.keys)) + " keys withheld>"
        return d


@dataclass
class Verdict:
    """The answer and -- always -- the reason. Refusal is the default."""
    granted: bool
    reason: str
    enumeration: str = ""
    enumeration_sha256: str = ""
    enumeration_retrieved_utc: str = ""
    query_key: str = ""
    evidence_kind: str = ""

    def __bool__(self) -> bool:      # so `if verdict:` cannot silently pass a refusal
        return self.granted


def earn_unobtainable(query_key: str, enumeration, evidence_kind: EvidenceKind) -> Verdict:
    """The ONLY route to GENUINELY_UNOBTAINABLE. Refuses by default, with a reason."""
    ek = str(getattr(evidence_kind, "value", evidence_kind))

    if enumeration is None:
        return Verdict(False, "REFUSED: no enumeration named. A failed fetch measures OUR "
                              "REACH, not the existence of the document.",
                       query_key=query_key, evidence_kind=ek)

    if evidence_kind not in VALID_EVIDENCE:
        return Verdict(
            False,
            "REFUSED: evidence_kind=" + repr(ek) + " measures OUR REACH, not the world. "
            "Only " + str(sorted(k.value for k in VALID_EVIDENCE)) + " can earn this state.",
            enumeration=enumeration.name, enumeration_sha256=enumeration.sha256,
            enumeration_retrieved_utc=enumeration.retrieved_utc,
            query_key=query_key, evidence_kind=ek)

    if not enumeration.sha256 or not enumeration.retrieved_utc:
        return Verdict(False, "REFUSED: enumeration carries no hash/retrieval date. A claim "
                              "about a register is a claim about a VERSION.",
                       enumeration=enumeration.name, query_key=query_key, evidence_kind=ek)

    if not enumeration.positive_control_passed:
        return Verdict(
            False,
            "REFUSED: positive control did not pass -- "
            + (enumeration.positive_control_note or "none was run")
            + ". An enumeration that can only answer 'absent' is not a check.",
            enumeration=enumeration.name, enumeration_sha256=enumeration.sha256,
            enumeration_retrieved_utc=enumeration.retrieved_utc,
            query_key=query_key, evidence_kind=ek)

    if enumeration.contains(query_key):
        return Verdict(
            False,
            "REFUSED: " + repr(query_key) + " IS a row in " + enumeration.name + ". The "
            "register says the document exists; failing to fetch it is a RETRIEVAL finding, "
            "not an evidence one.",
            enumeration=enumeration.name, enumeration_sha256=enumeration.sha256,
            enumeration_retrieved_utc=enumeration.retrieved_utc,
            query_key=query_key, evidence_kind=ek)

    return Verdict(
        True,
        "EARNED: " + repr(query_key) + " is absent from all " + str(enumeration.n_rows)
        + " rows of " + enumeration.name + " (sha256:" + enumeration.sha256[:12]
        + ", retrieved " + enumeration.retrieved_utc + "); positive control "
        + repr(enumeration.positive_control_key) + " was found in the same bytes. "
        + enumeration.absence_licenses,
        enumeration=enumeration.name, enumeration_sha256=enumeration.sha256,
        enumeration_retrieved_utc=enumeration.retrieved_utc,
        query_key=query_key, evidence_kind=ek)


# ---------------------------------------------------------------- EMA register
EMA_MEDICINES_XLSX = ("https://www.ema.europa.eu/en/documents/report/"
                      "medicines-output-medicines-report_en.xlsx")
EMA_ROW_MEANING = ("one row = one medicine EMA has assessed centrally. An EPAR exists "
                   "ONLY for a centrally authorised medicine.")
EMA_ABSENCE_LICENSES = ("Absence licenses exactly this: 'EMA has no centrally authorised "
                        "product under this name, therefore no EPAR exists.' It licenses "
                        "NOTHING about national authorisations, about FDA, or about the "
                        "trial literature.")


def ema_extract_names(payload: bytes) -> list:
    """Medicine names from the EMA xlsx. The header row is LOCATED, never assumed.

    Both the medicine name and the INN/active-substance column are harvested, because
    a product is reachable under either and a one-column read would manufacture
    absences.
    """
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    wanted = ("name of medicine", "medicine name", "name",
              "international non-proprietary name (inn) / common name",
              "international non-proprietary name/common name",
              "inn / common name", "active substance")
    cols: list = []
    names: list = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if not cols:
            low = [c.lower() for c in cells]
            cols = [i for i, c in enumerate(low) if c in wanted]
            continue
        for i in cols:
            if i < len(cells) and cells[i]:
                # An INN cell can hold several substances; split so each is a key.
                for part in str(cells[i]).replace(";", ",").split(","):
                    if part.strip():
                        names.append(part.strip())
    wb.close()
    return names


def build_ema_enumeration(positive_control: str = "Ferinject") -> Enumeration:
    """Fetch the live EMA register and pin it. Network required."""
    import requests
    ua = {"User-Agent": "oa68k-ladder/0.1 (mailto:mahmood726@gmail.com)"}
    r = requests.get(EMA_MEDICINES_XLSX, headers=ua, timeout=180)
    r.raise_for_status()
    return Enumeration.from_bytes(
        name="EMA medicines register (medicines-output-medicines-report_en.xlsx)",
        url=EMA_MEDICINES_XLSX, row_meaning=EMA_ROW_MEANING,
        absence_licenses=EMA_ABSENCE_LICENSES,
        payload=r.content, extract_keys=ema_extract_names,
        positive_control_key=positive_control)


# ------------------------------------------------------------------- selftest
def _selftest() -> int:
    """PLANT BOTH WAYS: watch it refuse on a defect, watch it grant on a clean case,
    restore, and assert the restoration. No network."""
    fails = []

    def check(label, cond):
        print(("  ok    " if cond else "  FAIL  ") + label)
        if not cond:
            fails.append(label)

    payload = b"alpha\nbravo\ncharlie\n"
    good = Enumeration.from_bytes(
        "toy register", "http://example/toy", "one row = one thing",
        "absence licenses only 'not in the toy register'.",
        payload, lambda b: b.decode().split(), positive_control_key="bravo")

    print("PLANT 1 -- clean case: absent key, control passed, rows as evidence")
    v = earn_unobtainable("delta", good, EvidenceKind.ENUMERATION_ROWS)
    check("granted for a genuinely absent key", v.granted)
    check("reason names the register AND its hash",
          "toy register" in v.reason and good.sha256[:12] in v.reason)

    print("PLANT 2 -- defect: a 404 offered as evidence of absence")
    v = earn_unobtainable("delta", good, EvidenceKind.HTTP_STATUS)
    check("REFUSED when the evidence is an http status", not v.granted)
    check("the refusal says it measures our reach", "REACH" in v.reason.upper())

    print("PLANT 3 -- defect: the key IS in the register")
    v = earn_unobtainable("bravo", good, EvidenceKind.ENUMERATION_ROWS)
    check("REFUSED when the register says it exists", not v.granted)

    print("PLANT 4 -- defect: positive control fails (empty/wrong register)")
    bad = Enumeration.from_bytes(
        "empty register", "http://example/empty", "one row = one thing", "",
        b"", lambda b: [], positive_control_key="bravo")
    check("an enumeration with 0 rows has a failed control", not bad.positive_control_passed)
    v = earn_unobtainable("delta", bad, EvidenceKind.ENUMERATION_ROWS)
    check("REFUSED when the control did not pass", not v.granted)

    print("PLANT 5 -- defect: no enumeration at all")
    v = earn_unobtainable("delta", None, EvidenceKind.ENUMERATION_ROWS)
    check("REFUSED with no enumeration", not v.granted)

    print("PLANT 6 -- defect: hash stripped, so the version is unknown")
    d = asdict(good)
    d["keys"] = good.keys
    d["sha256"] = ""
    v = earn_unobtainable("delta", Enumeration(**d), EvidenceKind.ENUMERATION_ROWS)
    check("REFUSED when the version is unknown", not v.granted)

    print("PLANT 7 -- defect: a search that returned nothing")
    v = earn_unobtainable("delta", good, EvidenceKind.SEARCH_MISS)
    check("REFUSED when the evidence is a search miss", not v.granted)

    print("RESTORE -- the clean case must still pass after every plant")
    v = earn_unobtainable("delta", good, EvidenceKind.ENUMERATION_ROWS)
    check("restoration asserted", v.granted)

    n = 10
    print("\nselftest: " + str(n - len(fails)) + "/" + str(n) + " assertions -- "
          + ("PASS" if not fails else "FAIL " + str(fails)))
    return 1 if fails else 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--ema", action="store_true", help="build the live EMA enumeration")
    ap.add_argument("--control", default="Ferinject", help="positive-control medicine name")
    ap.add_argument("--query", action="append", default=[],
                    help="name to test for absence (repeatable)")
    ap.add_argument("--out", default="")
    a = ap.parse_args()

    if a.selftest:
        return _selftest()

    if a.ema:
        enum = build_ema_enumeration(a.control)
        print(enum.name)
        print("  url     " + enum.url)
        print("  sha256  " + enum.sha256)
        print("  pulled  " + enum.retrieved_utc)
        print("  rows    " + str(enum.n_rows) + " distinct keys")
        print("  control " + enum.positive_control_note)
        if not enum.positive_control_passed:
            print("  ** positive control FAILED -- this register cannot grant absence **")
        out = {"enumeration": enum.summary(), "verdicts": []}
        for q in a.query:
            v = earn_unobtainable(q, enum, EvidenceKind.ENUMERATION_ROWS)
            print("\n  query " + repr(q))
            print("    granted=" + str(v.granted))
            print("    " + v.reason)
            out["verdicts"].append(asdict(v))
        if a.out:
            with open(a.out, "w", encoding="utf-8") as f:
                json.dump(out, f, indent=1)
            print("\nwrote " + a.out)
        return 0

    ap.print_help()
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
